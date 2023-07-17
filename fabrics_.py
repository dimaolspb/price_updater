from openpyxl import *
import shutil
import datetime
import os


''' наименования тканей переносим в буфер как есть, трансформируем (BLACK-OUT в б/о) при переносе в шаблон
цель буфера:

в нём хранятся данные, которые трудно извлекать из прайса, где они обозначены картинками,
поэтому в буфер добавляем данные в первый раз копированием, при этом создаём в исходном прайсе столбцы, в которых 
словами обозначаем смысл картинок.

При получении от поставщика нового прайса этот скрипт:
1. сравнивает позиции в прайсе и в буфере. Обновляет данные, но только строковые и числовые. Удаляет позиции, которые 
в новом прайсе больше не присутствуют. Добавляет новые, но заполняет только строковые и числовые данные, а те, что 
заполнены в прайсе картинками - нужно заполнять руками, иначе скрипт, переносящий потом данные в шаблон выдаст ошибку,
что не все данные заполнены'''


class FabricsPriceUpdate():
    def __init__(self):
        self.price_filename = 'price.xlsx'
        self.fabrics_sheet_name_in_price = 3
        self.buffer_filename = 'буфер.xlsx'
        self.fabrics_sheet_name_in_buffer = 'fabrics'

    def make_backup(self): # сделаем бэкап файла БУФЕР перед его изменением

        # Получаем текущую дату и время
        current_datetime = datetime.datetime.now()

        # Генерируем строку с текущей датой и временем в формате ГГГГММДД_ЧЧММСС
        timestamp = current_datetime.strftime("%Y%m%d_%H%M%S")

        # Получаем расширение файла
        file_extension = os.path.splitext(self.buffer_filename)[1]

        # Создаем новое имя файла с добавлением текущей даты и времени
        new_filename = f"{os.path.splitext(self.buffer_filename)[0]}_{timestamp}{file_extension}"

        # Копируем файл с новым именем
        shutil.copy(self.buffer_filename, new_filename)
        print('Бэкап файла БУФЕР создан')

    def make_actual_fabrics_list_in_buffer(self):# проверка на вхождение названий столбцов буфера в прайс
        print('В файле "прайс" включите видимость номеров колонок и строк (Вид / заголовки или показывать заголовки)')
        last_row_of_fabrics_in_price = int(input('Введите номер последней строки списка тканей в прайсе: '))
        print('Теперь закройте файлы price.xlsx и буфер.xlsx, после чего нажмите Enter')
        input()
        print('Идёт проверка соответствия названий столбцов в прайсе и в буфере')
        titles_row_in_price = 7 # номер строки прайса, где названия столбцов
        titles_row_in_buffer = 1 # номер строки буфера, где газвания столбцов

        wb_price = load_workbook(self.price_filename, data_only=True)
        sheet_in_price = wb_price.worksheets[self.fabrics_sheet_name_in_price]
        columns_titles_in_price = []
        for num in range(2, sheet_in_price.max_column): # создаём список названий колонок в прайсе
            if sheet_in_price[titles_row_in_price][num].value is not None:
                columns_titles_in_price.append(str(sheet_in_price[titles_row_in_price][num].value).lower())
        wb_buffer = load_workbook(self.buffer_filename, data_only=True)
        sheet_in_buffer = wb_buffer.worksheets[0]
        columns_titles_in_buffer = []
        for num in range(sheet_in_buffer.max_column): # создаём список названий колонок в буфере
            if sheet_in_buffer[titles_row_in_buffer][num].value is not None:
                columns_titles_in_buffer.append(str(sheet_in_buffer[titles_row_in_buffer][num].value).lower())

        #print('in price', columns_titles_in_price)
        #print('in buffer', columns_titles_in_buffer)

        if all(elem in columns_titles_in_price for elem in columns_titles_in_buffer):
            print("все заголовки колонок буфера содержатся в прайсе, можно обновлять список тканей")
            input('Для продолжения нажмите Enter')
        else:
            print("не все заголовки столбцов буфера содержаться в прайсе, приведите в соотвестсвие.", "В прайсе нет:",
                  sep='\n', end='\n')
            diff = set(columns_titles_in_buffer).difference(columns_titles_in_price)
            diff = list(diff)
            for i in range(len(diff)):
                print(diff[i])
            return

    # определим какие позиции удалить в БУФЕРЕ, а какие в него добавить

        first_row_of_fabrics_in_price = titles_row_in_price + 1
        first_row_of_fabrics_in_buffer = titles_row_in_buffer + 1
        column_num_with_fabric_titles_in_price = 2
        column_num_fabrics_titles_in_buffer = 0
        fabrics_in_price = set()
        fabrics_in_buffer = set()

        for row in range(first_row_of_fabrics_in_price, last_row_of_fabrics_in_price + 1): # наполняем множество тканей прайса
            fabrics_in_price.add(sheet_in_price[row][column_num_with_fabric_titles_in_price].value)

        for row in range(first_row_of_fabrics_in_buffer, sheet_in_buffer.max_row + 1): # наполняем множество тканей буфера
            fabrics_in_buffer.add(sheet_in_buffer[row][column_num_fabrics_titles_in_buffer].value)
        fabrics_in_buffer = {x for x in fabrics_in_buffer if x is not None} # удаляем все None

        to_delete = list(fabrics_in_buffer - fabrics_in_price)
        to_add = list(fabrics_in_price - fabrics_in_buffer)

        fabrics_in_buffer = list(fabrics_in_buffer)
        fabrics_in_buffer.sort()

        print(f'Удалить из БУФЕРА: {to_delete}')
        print(f'Добавить в БУФЕР: {to_add}')

        fabrics_in_buffer = [item for item in fabrics_in_buffer if item not in to_delete]
        for i in range(len(fabrics_in_buffer)):
            print(i, fabrics_in_buffer[i])

        fabrics_in_price = list(fabrics_in_price)
        fabrics_in_price.sort()
        print('from price')
        for i in range(len(fabrics_in_price)):
            print(i, fabrics_in_price[i])

        # теперь удалим лишние позиции




        # теперь добавим новые позиции

        # теперь отсортируем список

        # # Загрузка файла
        # df = pd.read_excel('имя_файла.xlsx')
        #
        # # Сортировка данных по первому столбцу
        # sorted_df = df.sort_values(by='название_столбца_для_сортировки')
        #
        # # Сохранение изменений в файле
        # sorted_df.to_excel('имя_файла.xlsx', index=False)

fabrics_price_update = FabricsPriceUpdate()
fabrics_price_update.make_backup()
fabrics_price_update.make_actual_fabrics_list_in_buffer()

